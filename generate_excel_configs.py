import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = r"d:\utest\FinalRPG\Assets\Data\ExcelConfig"
os.makedirs(OUT_DIR, exist_ok=True)

HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
NOTE_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
NOTE_FONT = Font(color="595959", italic=True, size=9)


def set_header_row(ws, headers, col_widths=None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20


def write_note_row(ws, note, ncols):
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=note).font = NOTE_FONT
    ws.cell(row=r, column=1).fill = NOTE_FILL
    for ci in range(2, ncols + 1):
        ws.cell(row=r, column=ci).fill = NOTE_FILL


# -----------------------------------------------------------------------
# 1. ItemsConfig.xlsx
# -----------------------------------------------------------------------
ITEM_HEADERS = [
    "assetName", "itemID", "itemName", "description",
    "iconPath", "category", "rarity", "isStackable", "maxStack"
]
ITEM_WIDTHS = [38, 20, 22, 60, 40, 14, 12, 12, 11]

ITEMS_DATA = [
    ("CONS_BUFF_002_AdvancedStrengthPotion", "CONS_BUFF_002",
     u"\u9ad8\u7ea7\u529b\u91cf\u836f\u5242",
     u"\u70bc\u91d1\u5e08\u8c03\u914d\u7684\u5f3a\u6548\u589e\u76ca\u836f\u5242\uff0c\u996e\u7528\u540e30\u5206\u949f\u5185\u63d0\u5347"
     u"20%\u7269\u7406\u653b\u51fb\u529b\uff0c\u6b7b\u4ea1\u540e\u6548\u679c\u4e0d\u6d88\u5931\u3002",
     "PackageIcon/Consumable/CONS_BUFF_002", "Consumable", "Uncommon", "TRUE", 99),

    ("CONS_BUFF_003_AllAttributePotion", "CONS_BUFF_003",
     u"\u5168\u5c5e\u6027\u589e\u5e45\u79d8\u836f",
     u"\u53f2\u8bd7\u7ea7\u70bc\u91d1\u79d8\u836f\uff0c\u996e\u7528\u540e1\u5c0f\u65f6\u5185\u5168\u5c5e\u6027\u63d0\u5347"
     u"15%\uff0c\u83b7\u5f9710%\u4f24\u5bb3\u51cf\u514d\uff0c\u4ec5\u53ef\u5728\u975e\u6218\u6597\u72b6\u6001\u4f7f\u7528\u3002",
     "PackageIcon/Consumable/CONS_BUFF_003", "Consumable", "Epic", "TRUE", 20),

    ("CONS_HP_001_SmallHealthPotion", "CONS_HP_001",
     u"\u5c0f\u578b\u6cbb\u7597\u836f\u6c34",
     u"\u5192\u9669\u8005\u51fa\u95e8\u5fc5\u5907\u7684\u57fa\u7840\u836f\u6c34\uff0c\u996e\u7528\u540e\u7acb\u5373\u6062\u590d200\u70b9\u751f\u547d\u503c\u3002",
     "PackageIcon/Consumable/CONS_HP_001", "Consumable", "Common", "TRUE", 99),

    ("EQ_ACC_003_EternalRing", "EQ_ACC_003",
     u"\u6c38\u71c3\u4e4b\u6212",
     u"\u9576\u5d4c\u706b\u7130\u6838\u5fc3\u7684\u4f20\u5947\u6212\u6307\uff0c\u4f69\u6234\u540e\u6c38\u4e45\u63d0\u5347"
     u"15%\u706b\u7130\u6cd5\u672f\u4f24\u5bb3\uff0c\u653b\u51fb\u65f6\u6709\u6982\u7387\u89e6\u53d1\u8303\u56f4\u706b\u7130\u7206\u70b8\u3002",
     "PackageIcon/Equipment/EQ_ACC_003", "Equipment", "Legendary", "FALSE", 1),

    ("EQ_ARMOR_002_PlateArmor", "EQ_ARMOR_002",
     u"\u5b88\u62a4\u4e4b\u677f\u7532\u80f8\u7532",
     u"\u77ee\u4eba\u94c1\u5320\u6253\u9020\u7684\u91cd\u578b\u677f\u7532\uff0c\u8868\u9762\u523b\u6709\u9632\u5fa1\u7b26\u6587\uff0c"
     u"\u5927\u5e45\u63d0\u5347\u7269\u7406\u6297\u6027\uff0c\u964d\u4f4e\u9504\u5668\u4e0e\u9b27\u5668\u4f24\u5bb3\u3002",
     "PackageIcon/Equipment/EQ_ARMOR_002", "Equipment", "Rare", "FALSE", 1),

    ("EQ_WEAP_001_IronSword", "EQ_WEAP_001",
     u"\u94c1\u5236\u957f\u5251",
     u"\u5192\u9669\u8005\u5165\u95e8\u6807\u914d\u5355\u624b\u5251\uff0c\u719f\u94c1\u953b\u9020\uff0c\u5e73\u8861\u6027\u5c1a\u53ef\uff0c"
     u"\u80fd\u5e94\u5bf9\u91ce\u5916\u591a\u6570\u4f4e\u9636\u9b54\u7269\u3002",
     "PackageIcon/Equipment/EQ_WEAP_001", "Equipment", "Common", "FALSE", 1),

    ("ITEM_BAG_003_BagExpansionStone", "ITEM_BAG_003",
     u"\u80cc\u5305\u6269\u5bb9\u77f3",
     u"\u8574\u542b\u7a7a\u95f4\u9b54\u6cd5\u7684\u7a00\u6709\u6676\u77f3\uff0c\u4f7f\u7528\u540e\u6c38\u4e45\u89e3\u9501"
     u"8\u4e2a\u80cc\u5305\u683c\u5b50\uff0c\u6bcf\u4e2a\u89d2\u8272\u6700\u591a\u53ef\u4f7f\u752810\u6b21\u3002",
     "PackageIcon/Item/ITEM_BAG_003", "Item", "Epic", "TRUE", 10),

    ("ITEM_KEY_001_CopperKey", "ITEM_KEY_001",
     u"\u94dc\u5236\u5b9d\u7bb1\u9470\u5319",
     u"\u7528\u4e8e\u5f00\u542f\u91ce\u5916\u5e38\u89c1\u94dc\u5236\u5b9d\u7bb1\u7684\u9470\u5319\uff0c\u5b9d\u7bb1\u5185\u85cf\u6709"
     u"\u5192\u9669\u8005\u7559\u4e0b\u7684\u7269\u8d44\u4e0e\u88c5\u5907\u3002",
     "PackageIcon/Item/ITEM_KEY_001", "Item", "Common", "TRUE", 99),

    ("ITEM_TP_002_TeleportScroll", "",
     u"\u4e3b\u57ce\u4f20\u9001\u5377\u8f74",
     u"\u9644\u9b54\u5e08\u5236\u4f5c\u7684\u4f20\u9001\u5377\u8f74\uff0c\u4f7f\u7528\u540e\u7acb\u5373\u4f20\u9001\u56de\u4e3b\u57ce"
     u"\u5b89\u5168\u533a\uff0c\u662f\u91ce\u5916\u9047\u9669\u7684\u7edd\u4f73\u9003\u751f\u9053\u5177\u3002",
     "", "Item", "Uncommon", "TRUE", 20),

    ("MAT_ENCHANT_002_EnchantCrystal", "MAT_ENCHANT_002",
     u"\u9644\u9b54\u6c34\u6676",
     u"\u8574\u542b\u7eaf\u51c0\u9b54\u6cd5\u80fd\u91cf\u7684\u7a00\u6709\u6c34\u6676\uff0c\u88c5\u5907\u9644\u9b54\u3001"
     u"\u7b26\u6587\u94ed\u523b\u7684\u5fc5\u5907\u6750\u6599\uff0c\u54c1\u8d28\u8d8a\u9ad8\u9644\u9b54\u6210\u529f\u7387\u8d8a\u9ad8\u3002",
     "PackageIcon/Material/MAT_ENCHANT_002", "Material", "Rare", "TRUE", 99),

    ("MAT_LEGEND_003_DragonScale", "MAT_LEGEND_003",
     u"\u9f99\u9cde\u788e\u7247",
     u"\u8fdc\u53e4\u5de8\u9f99\u8131\u843d\u7684\u9cde\u7247\u788e\u7247\uff0c\u8574\u542b\u6781\u5f3a\u7684\u5de8\u9f99\u4e4b\u529b\uff0c"
     u"\u6253\u9020\u4f20\u8bf4\u7ea7\u88c5\u5907\u7684\u6838\u5fc3\u6750\u6599\uff0c\u6781\u4e3a\u7a00\u6709\u3002",
     "PackageIcon/Material/MAT_LEGEND_003", "Material", "Legendary", "TRUE", 50),

    ("MAT_ORE_001_IronOre", "MAT_ORE_001",
     u"\u94c1\u77ff\u77f3",
     u"\u6700\u5e38\u89c1\u7684\u91d1\u5c5e\u77ff\u77f3\uff0c\u953b\u9020\u57fa\u7840\u6b66\u5668\u4e0e\u62a4\u7532\u7684\u6838\u5fc3\u6750\u6599\uff0c"
     u"\u91ce\u5916\u77ff\u8109\u968f\u5904\u53ef\u91c7\u96c6\u3002",
     "PackageIcon/Material/MAT_ORE_001", "Material", "Common", "TRUE", 999),

    ("OTHER_COLLECT_002_AnniversaryCoin", "OTHER_COLLECT_002",
     u"\u5192\u9669\u8005\u5468\u5e74\u7eaa\u5ff5\u5e01",
     u"\u4e3b\u57ce\u5192\u9669\u8005\u534f\u4f1a\u53d1\u884c\u7684\u9650\u91cf\u7eaa\u5ff5\u5e01\uff0c\u65e0\u5b9e\u9645\u529f\u80fd\uff0c"
     u"\u662f\u8d44\u6df1\u5192\u9669\u8005\u7684\u8eab\u4efd\u8c61\u5f81\uff0c\u6781\u5177\u6536\u85cf\u4ef7\u5024\u3002",
     "PackageIcon/Other/OTHER_COLLECT_002", "Other", "Rare", "FALSE", 1),

    ("OTHER_FUN_003_SlimeMask", "OTHER_FUN_003",
     u"\u6076\u641e\u53f2\u83b1\u59c6\u9762\u5177",
     u"\u7528\u53f2\u83b1\u59c6\u51dd\u80f6\u5236\u4f5c\u7684\u8da3\u5473\u9762\u5177\uff0c\u4f69\u6234\u540e\u89d2\u8272\u5916\u89c2"
     u"\u4f1a\u53d8\u6210\u53f2\u83b1\u59c6\uff0c\u4ec5\u4f5c\u5a31\u4e50\u4f7f\u7528\uff0c\u65e0\u4efb\u4f55\u5c5e\u6027\u52a0\u6210\u3002",
     "PackageIcon/Other/OTHER_FUN_003", "Other", "Uncommon", "FALSE", 1),

    ("OTHER_JUNK_001_BrokenBottle", "OTHER_JUNK_001",
     u"\u7834\u635f\u7684\u7a7a\u9152\u74f6",
     u"\u5192\u9669\u8005\u4e22\u5f03\u7684\u7a7a\u9152\u74f6\uff0c\u65e0\u4efb\u4f55\u5b9e\u9645\u7528\u9014\uff0c"
     u"\u53ea\u80fd\u5356\u7ed9\u5546\u4eba\u6362\u53d6\u5c11\u91cf\u91d1\u5e01\u3002",
     "PackageIcon/Other/OTHER_JUNK_001", "Other", "Common", "TRUE", 99),

    ("QUEST_001_VillageLetter", "QUEST_001",
     u"\u6751\u957f\u7684\u5bb6\u4e66",
     u"\u65b0\u624b\u6751\u6751\u957f\u6258\u4ed8\u7684\u5bb6\u4e66\uff0c\u9700\u8f6c\u4ea4\u7ed9\u4e3b\u57ce\u9a7b\u5b88\u7684\u6751\u957f"
     u"\u513f\u5b50\uff0c\u76d6\u6709\u65b0\u624b\u6751\u4e13\u5c5e\u5370\u7ae0\u3002",
     "PackageIcon/QuestItem/QUEST_001", "QuestItem", "Common", "FALSE", 1),

    ("QUEST_002_AncientStoneTablet", "QUEST_002",
     u"\u7834\u635f\u7684\u53e4\u4ee3\u77f3\u677f",
     u"\u53e4\u4ee3\u9057\u8403\u4e2d\u51fa\u571f\u7684\u77f3\u677f\uff0c\u523b\u6709\u65e0\u4eba\u80fd\u61c2\u7684\u53e4\u6587\u5b57\uff0c"
     u"\u8003\u53e4\u5b66\u5bb6\u6b63\u5728\u5bfb\u627e\u5b8c\u6574\u77f3\u677f\u89e3\u8bfb\u79d8\u5bc6\u3002",
     "PackageIcon/QuestItem/QUEST_002", "QuestItem", "Uncommon", "FALSE", 1),

    ("QUEST_003_DemonHeartFragment", "QUEST_003",
     u"\u9b54\u738b\u7684\u5fc3\u810f\u788e\u7247",
     u"\u51fb\u8d25\u9b54\u738b\u5206\u8eab\u6389\u843d\u7684\u6838\u5fc3\u788e\u7247\uff0c\u8574\u542b\u5fae\u5f31\u6df1\u6e0a\u4e4b\u529b\uff0c"
     u"\u662f\u5c01\u5370\u9b54\u738b\u7684\u5173\u952e\u9053\u5177\uff0c\u7edd\u4e0d\u80fd\u843d\u5165\u6076\u4eba\u624b\u4e2d\u3002",
     "PackageIcon/QuestItem/QUEST_003", "QuestItem", "Legendary", "FALSE", 1),
]

wb_items = openpyxl.Workbook()
ws_items = wb_items.active
ws_items.title = "Items"
set_header_row(ws_items, ITEM_HEADERS, ITEM_WIDTHS)
ws_items.freeze_panes = "A2"

for row in ITEMS_DATA:
    ws_items.append(list(row))
    ws_items.cell(row=ws_items.max_row, column=4).alignment = Alignment(wrap_text=True)
    ws_items.row_dimensions[ws_items.max_row].height = 30

write_note_row(ws_items,
    "# category: Item/Equipment/QuestItem/Consumable/Material/Other  |  "
    "rarity: Common/Uncommon/Rare/Epic/Legendary  |  isStackable: TRUE/FALSE",
    len(ITEM_HEADERS))

path_items = os.path.join(OUT_DIR, "ItemsConfig.xlsx")
wb_items.save(path_items)
print("OK ItemsConfig.xlsx (%d rows) -> %s" % (len(ITEMS_DATA), path_items))


# -----------------------------------------------------------------------
# 2. QuestsConfig.xlsx  (3 sheets)
# targetType: Kill / Collect / Communicate
# rewardType: Item / Currency / Experience
# -----------------------------------------------------------------------
QUESTS_DATA = [
    ("ClearBlackForest_Phase1_Investigate",
     "ClearBlackForest_Phase1_Investigate",
     u"\u8c03\u67e5\u9b54\u7269\u771f\u76f8",
     u"\u6751\u5b50\u897f\u8fb9\u7684\u9ed1\u98ce\u6797\u51fa\u73b0\u4e86\u795e\u79d8\u9b54\u7269\uff0c"
     u"\u6751\u6c11\u65e0\u6cd5\u7838\u67f4\uff0c\u5b58\u7cae\u5373\u5c06\u8017\u5c3d\u3002\u9700\u8981\u8c03\u67e5\u771f\u76f8\u5e76\u51b3\u5b9a\u5bf9\u7b56\u3002",
     1),
    ("ClearBlackForest_Phase2_Kill",
     "ClearBlackForest_Phase2_Kill",
     u"\u7ec8\u7ed3\u9ed1\u98ce\u6797\u5371\u673a",
     u"\u6751\u957f\u51b3\u5b9a\u4ee5\u6b66\u529b\u89e3\u51b3\u9ed1\u98ce\u6797\u7684\u9b54\u7269\u5a01\u80c1\u3002"
     u"\u518d\u6b21\u524d\u5f80\u9ed1\u98ce\u6797\uff0c\u6d88\u706d\u72c2\u6012\u7684\u9b54\u7269\u3002",
     1),
    ("ClearBlackForest_Phase2_Peace",
     "ClearBlackForest_Phase2_Peace",
     u"\u6cbb\u6108\u9ed1\u98ce\u6797\u4e4b\u4e3b",
     u"\u6751\u957f\u51b3\u5b9a\u5c1d\u8bd5\u548c\u5e73\u89e3\u51b3\u3002"
     u"\u4f7f\u7528\u6536\u96c6\u5230\u7684\u8349\u836f\u5b89\u629a\u72c2\u66b4\u7684\u9b54\u7269\uff0c\u5316\u89e3\u5371\u673a\u3002",
     1),
]

OBJECTIVES_DATA = [
    ("ClearBlackForest_Phase1_Investigate", 1, "Communicate", "VillageChief", 1,
     u"\u4e0e\u6751\u957f\u4ea4\u8c08\uff0c\u4e86\u89e3\u9ed1\u98ce\u6797\u60c5\u51b5"),
    ("ClearBlackForest_Phase1_Investigate", 2, "Collect", "Clue_Footprint", 1,
     u"\u524d\u5f80\u9ed1\u98ce\u6797\u5916\u56f4\uff0c\u5bfb\u627e\u811a\u5370\u7b49\u7ebf\u7d22"),
    ("ClearBlackForest_Phase1_Investigate", 3, "Collect", "Clue_Herb", 1,
     u"\u6df1\u5165\u6797\u4e2d\uff0c\u6536\u96c6\u88ab\u67d0\u79cd\u751f\u7269\u5543\u98df\u7684\u8349\u836f"),
    ("ClearBlackForest_Phase1_Investigate", 4, "Communicate", "VillageChief", 1,
     u"\u5e26\u56de\u8bc1\u636e\u5411\u6751\u957f\u62a5\u544a\uff0c\u5e76\u51b3\u5b9a\u5bf9\u7b56"),
    ("ClearBlackForest_Phase2_Kill", 1, "Kill", "ForestMonster", 1,
     u"\u518d\u6b21\u524d\u5f80\u9ed1\u98ce\u6797\uff0c\u6d88\u706d\u72c2\u6012\u7684\u9b54\u7269"),
    ("ClearBlackForest_Phase2_Kill", 2, "Communicate", "VillageChief", 1,
     u"\u5411\u6751\u957f\u62a5\u544a\u9b54\u7269\u5df2\u88ab\u6d88\u706d"),
    ("ClearBlackForest_Phase2_Peace", 1, "Communicate", "Monster_BlackForestBoss", 1,
     u"\u9760\u8fd1\u9b54\u7269\uff0c\u4f7f\u7528\u8349\u836f\u5b89\u629a\u5b83"),
    ("ClearBlackForest_Phase2_Peace", 2, "Communicate", "VillageChief", 1,
     u"\u5411\u6751\u957f\u62a5\u544a\u4e8b\u60c5\u5df2\u548c\u5e73\u89e3\u51b3"),
]

REWARDS_DATA = [
    ("ClearBlackForest_Phase2_Kill",  1, "Currency",   "",              200),
    ("ClearBlackForest_Phase2_Kill",  2, "Experience", "",              500),
    ("ClearBlackForest_Phase2_Peace", 1, "Currency",   "",              500),
    ("ClearBlackForest_Phase2_Peace", 2, "Experience", "",              600),
    ("ClearBlackForest_Phase2_Peace", 3, "Item",       "Item_RareHerb", 3),
]

wb_quests = openpyxl.Workbook()

ws_q = wb_quests.active
ws_q.title = "Quests"
q_headers = ["assetName", "id", "title", "description", "isOrdered"]
q_widths = [45, 45, 20, 65, 11]
set_header_row(ws_q, q_headers, q_widths)
ws_q.freeze_panes = "A2"
for row in QUESTS_DATA:
    ws_q.append(list(row))
    ws_q.cell(ws_q.max_row, 4).alignment = Alignment(wrap_text=True)
    ws_q.row_dimensions[ws_q.max_row].height = 40
write_note_row(ws_q, "# isOrdered: 1=ordered, 0=unordered", len(q_headers))

ws_o = wb_quests.create_sheet("Objectives")
o_headers = ["assetName", "idx", "targetType", "targetId", "requiredAmount", "uiDescription"]
o_widths = [45, 6, 14, 28, 15, 40]
set_header_row(ws_o, o_headers, o_widths)
ws_o.freeze_panes = "A2"
for row in OBJECTIVES_DATA:
    ws_o.append(list(row))
write_note_row(ws_o, "# targetType: Kill / Collect / Communicate", len(o_headers))

ws_r = wb_quests.create_sheet("Rewards")
r_headers = ["assetName", "idx", "rewardType", "rewardId", "amount"]
r_widths = [45, 6, 14, 25, 10]
set_header_row(ws_r, r_headers, r_widths)
ws_r.freeze_panes = "A2"
for row in REWARDS_DATA:
    ws_r.append(list(row))
write_note_row(ws_r, "# rewardType: Item / Currency / Experience", len(r_headers))

path_quests = os.path.join(OUT_DIR, "QuestsConfig.xlsx")
wb_quests.save(path_quests)
print("OK QuestsConfig.xlsx (%d quests, %d objectives, %d rewards) -> %s" % (
    len(QUESTS_DATA), len(OBJECTIVES_DATA), len(REWARDS_DATA), path_quests))
print("Done!")
