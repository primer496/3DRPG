import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = r"d:\utest\FinalRPG\Assets\Data\ExcelConfig"
os.makedirs(OUT_DIR, exist_ok=True)

HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
NOTE_FONT = Font(color="595959", italic=True, size=9)
ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")


def set_header_row(ws, headers, col_widths=None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18


def append_data_row(ws, row_data, use_alt=False):
    ws.append(list(row_data))
    if use_alt:
        r = ws.max_row
        for ci in range(1, len(row_data) + 1):
            ws.cell(r, ci).fill = ALT_FILL


def write_note_row(ws, note, ncols):
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="# " + note).font = NOTE_FONT
    ws.cell(row=r, column=1).fill = NOTE_FILL
    for ci in range(2, ncols + 1):
        ws.cell(row=r, column=ci).fill = NOTE_FILL


# -----------------------------------------------------------------------
# CharacterConfig.xlsx - 5 sheets
# assetPath: relative to Assets/ (no Assets/ prefix, no .asset suffix)
# -----------------------------------------------------------------------
wb = openpyxl.Workbook()

# ===== Sheet 1: Movement =====
ws_mv = wb.active
ws_mv.title = "Movement"
mv_headers = [
    "assetPath",
    "moveSpeed", "accel", "runSpeedMultiplier",
    "enableStopState",
    "stopDuration", "stopEnterSpeedThreshold", "stopEnterCrossFade", "stopSpeedDecayTime"
]
mv_widths = [55, 12, 10, 20, 15, 14, 24, 20, 18]
set_header_row(ws_mv, mv_headers, mv_widths)
ws_mv.freeze_panes = "B2"

MV_DATA = [
    # assetPath, moveSpeed, accel, runSpeedMultiplier, enableStopState,
    # stopDuration, stopEnterSpeedThreshold, stopEnterCrossFade, stopSpeedDecayTime
    ("Resources/GameConfigs/Character/Modules/Player/ActorMovementConfig",
     1, 40, 2, "TRUE", 0.74, 0.63, 0.11, 0.23),
    ("Resources/GameConfigs/Character/Modules/Enemy/ActorMovementConfig",
     1, 40, 2, "TRUE", 0.26, 0.32, 0.06, 0.12),
]
for i, row in enumerate(MV_DATA):
    append_data_row(ws_mv, row, use_alt=(i % 2 == 1))
write_note_row(ws_mv, "enableStopState: TRUE/FALSE  |  assetPath relative to Assets/ without .asset", len(mv_headers))


# ===== Sheet 2: Jump =====
ws_jp = wb.create_sheet("Jump")
jp_headers = ["assetPath", "jumpHeight", "jumpSpeed", "jumpGroundDetachTime"]
jp_widths = [55, 12, 12, 22]
set_header_row(ws_jp, jp_headers, jp_widths)
ws_jp.freeze_panes = "B2"

JP_DATA = [
    ("Resources/GameConfigs/Character/Modules/Player/ActorJumpConfig",
     1, 7, 0.12),
]
for i, row in enumerate(JP_DATA):
    append_data_row(ws_jp, row, use_alt=(i % 2 == 1))
write_note_row(ws_jp, "Player only - no Enemy jump config", len(jp_headers))


# ===== Sheet 3: Combat =====
ws_cb = wb.create_sheet("Combat")
cb_headers = [
    "assetPath",
    "attackRange", "comboResetTime", "maxComboSteps",
    "useCombatRootMotion", "combatRootMotionPlanarScale",
    "comboExitNormalizedTime", "hitReactionExitNormalizedTime",
    "hitKnockbackSpeed", "hitKnockbackDecay", "hitStopDuration", "hitStopRootMotionScale",
    "aimAssistRadius", "aimAssistAngle",
    "combo1WindowStart", "combo1WindowEnd",
    "combo2WindowStart", "combo2WindowEnd",
    "combo3WindowStart", "combo3WindowEnd",
    "combo4WindowStart", "combo4WindowEnd",
]
cb_widths = [55, 12, 14, 14, 20, 26,
             24, 28,
             18, 18, 16, 22,
             14, 14,
             18, 16, 18, 16, 18, 16, 18, 16]
set_header_row(ws_cb, cb_headers, cb_widths)
ws_cb.freeze_panes = "B2"

CB_DATA = [
    # Player
    ("Resources/GameConfigs/Character/Modules/Player/ActorCombatConfig",
     1.5, 0.6, 4, "TRUE", 1.0,
     0.95, 0.1,
     4.0, 15.0, 0.15, 0.1,
     6.0, 180.0,
     0.5, 0.755, 0.378, 0.764, 0.384, 0.66, 0.628, 0.976),
    # Enemy
    ("Resources/GameConfigs/Character/Modules/Enemy/ActorCombatConfig",
     0.5, 0.6, 4, "TRUE", 0.5,
     0.95, 0.58,
     4.0, 15.0, 0.15, 0.1,
     6.0, 180.0,
     0.38, 0.7, 0.22, 0.68, 0.2, 0.66, 0.18, 0.6),
]
for i, row in enumerate(CB_DATA):
    append_data_row(ws_cb, row, use_alt=(i % 2 == 1))
write_note_row(ws_cb, "useCombatRootMotion: TRUE/FALSE  |  hitKnockback/hitStop/aimAssist use C# defaults if not in YAML", len(cb_headers))


# ===== Sheet 4: Traversal =====
ws_tr = wb.create_sheet("Traversal")
tr_headers = [
    "assetPath",
    "vaultDuration", "vaultEnterCrossFade", "vaultExitCrossFade", "vaultExitNormalizedTime",
    "vaultLateDownStartNormalizedTime", "vaultLateDownSpeed", "vaultMinMoveSpeed",
    "vaultWallMask",
    "vaultDetectDistance", "vaultMaxFacingAngle",
    "vaultMinHeight", "vaultMaxHeight",
    "vaultSampleMinHeight", "vaultSampleMaxHeight", "vaultHeightSamples", "vaultDebugLog",
    "climbWallMask",
    "climbDetectDistance", "climbMaxFacingAngle",
    "climbSampleMinHeight", "climbSampleMaxHeight", "climbHeightSamples",
    "climbEnterCrossFade", "climbExitCrossFade", "climbExitNormalizedTime",
    "climb17ExitNormalizedTime", "climb17PlanarAssistSpeed", "climb17MinPlanarSpeed", "climbDebugLog",
    "wallActionAlignDuration", "wallActionAlignMinAngle",
    "vaultReferenceWallHeight", "climb05ReferenceWallHeight", "climb10ReferenceWallHeight",
    "climb17ReferenceWallHeight", "climb20ReferenceWallHeight",
    "wallActionHeightAdjustSpeed", "wallActionMaxUpOffset", "wallActionMaxDownOffset",
]
# Set column widths - first wide, rest narrow
tr_widths = [55] + [12] * (len(tr_headers) - 1)
tr_widths[1] = 14   # vaultDuration
tr_widths[8] = 14   # vaultWallMask
tr_widths[17] = 14  # climbWallMask
set_header_row(ws_tr, tr_headers, tr_widths)
ws_tr.freeze_panes = "B2"

TR_DATA = [
    ("Resources/GameConfigs/Character/Modules/Player/ActorTraversalConfig",
     0.42, 0.08, 0.1, 0.9,
     0.27, 3.22, 0.2,
     128,   # vaultWallMask (LayerMask m_Bits)
     0.75, 45,
     0.75, 1.2,
     0.2, 1.6, 6, "FALSE",
     128,   # climbWallMask (LayerMask m_Bits)
     0.75, 45,
     0.1, 2.4, 10,
     0.1, 0.12, 0.92,
     0.975, 0.45, 0.08, "FALSE",
     0.08, 8,
     1, 0.5, 1,
     1.5, 2,
     2.4, 0.6, 0.45),
]
for i, row in enumerate(TR_DATA):
    append_data_row(ws_tr, row, use_alt=(i % 2 == 1))
write_note_row(ws_tr, "vaultWallMask / climbWallMask: integer (LayerMask.value, e.g. 128 = layer 7)  |  DebugLog: TRUE/FALSE", len(tr_headers))


# ===== Sheet 5: ConfigSets =====
ws_cs = wb.create_sheet("ConfigSets")
cs_headers = [
    "assetPath", "soType",
    "enableLocomotion", "enableCombat", "enableJump", "enableTraversal",
]
cs_widths = [60, 10, 18, 14, 12, 16]
set_header_row(ws_cs, cs_headers, cs_widths)
ws_cs.freeze_panes = "B2"

CS_DATA = [
    ("Resources/GameConfigs/Character/Sets/Player/PlayerCapabilityConfigSet",
     "Player", "TRUE", "TRUE", "TRUE", "TRUE"),
    ("Resources/GameConfigs/Character/Sets/Enemy/EnemyCapabilityConfigSet",
     "Enemy", "TRUE", "TRUE", "FALSE", "FALSE"),
]
for i, row in enumerate(CS_DATA):
    append_data_row(ws_cs, row, use_alt=(i % 2 == 1))
write_note_row(ws_cs, "soType: Player or Enemy  |  enable*: TRUE/FALSE  |  SO references (movement/combat/jump/traversal) are NOT managed here", len(cs_headers))


path_out = os.path.join(OUT_DIR, "CharacterConfig.xlsx")
wb.save(path_out)
print("OK CharacterConfig.xlsx -> %s" % path_out)
print("  Sheets: Movement(%d), Jump(%d), Combat(%d), Traversal(%d), ConfigSets(%d)" % (
    len(MV_DATA), len(JP_DATA), len(CB_DATA), len(TR_DATA), len(CS_DATA)))
print("Done!")
