import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Items"

headers = ["itemID", "itemName", "description", "iconPath", "category", "rarity", "isStackable", "maxStack"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data = [
    ["ITEM_001", "铁剑", "一把普通的铁剑，适合初学者使用。", "Icons/iron_sword", "Equipment", "Common", "FALSE", 1],
    ["ITEM_002", "生命药水", "恢复100点生命值的红色药水。", "Icons/health_potion", "Consumable", "Common", "TRUE", 99],
    ["ITEM_003", "铁矿石", "冶炼铁制装备所需的原材料。", "Icons/iron_ore", "Material", "Common", "TRUE", 99],
    ["ITEM_004", "传说之剑", "据说由神明亲手锻造的神器。", "Icons/legend_sword", "Equipment", "Legendary", "FALSE", 1],
    ["ITEM_005", "任务卷轴", "记载着神秘任务线索的古老卷轴。", "Icons/quest_scroll", "QuestItem", "Uncommon", "FALSE", 1],
]

for row_data in data:
    ws.append(row_data)

col_widths = [12, 15, 35, 25, 12, 12, 12, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(r"d:\utest\FinalRPG\Assets\Data\ExcelConfig\ItemsConfig.xlsx")
print("Excel file created successfully")
