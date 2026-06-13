#!/usr/bin/env python3
"""Excel -> CSV 中间层转换脚本
用法: python excel_to_csv.py
将 Assets/Data/ExcelConfig/ 下所有 .xlsx 文件转换为 Assets/Data/CSV/ 下的 .csv 文件
"""
import os
import csv
import openpyxl

EXCEL_DIR = r"d:\utest\FinalRPG\Assets\Data\ExcelConfig"
CSV_DIR   = r"d:\utest\FinalRPG\Assets\Data\CSV"

def excel_to_csv(xlsx_path: str, csv_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 如果只有一个sheet，csv文件名与xlsx同名；多sheet则加sheet名后缀
        if len(wb.sheetnames) == 1:
            out_path = csv_path
        else:
            base, ext = os.path.splitext(csv_path)
            out_path = f"{base}_{sheet_name}{ext}"
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                # 跳过全空行
                if all(v is None for v in row):
                    continue
                writer.writerow([("" if v is None else str(v)) for v in row])
        print(f"  -> {out_path}")

def main():
    os.makedirs(CSV_DIR, exist_ok=True)
    xlsx_files = [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx")]
    if not xlsx_files:
        print("未找到xlsx文件")
        return
    for xlsx_file in xlsx_files:
        xlsx_path = os.path.join(EXCEL_DIR, xlsx_file)
        csv_name  = os.path.splitext(xlsx_file)[0] + ".csv"
        csv_path  = os.path.join(CSV_DIR, csv_name)
        print(f"转换: {xlsx_file}")
        excel_to_csv(xlsx_path, csv_path)
    print("完成！")

if __name__ == "__main__":
    main()
